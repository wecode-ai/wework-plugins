#!/usr/bin/env python3
# SPDX-License-Identifier: Apache-2.0
"""Deterministic XLSX creation, inspection, editing, rendering, and validation."""

from __future__ import annotations

import argparse
import csv
import json
import os
import shutil
import subprocess
import sys
import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import range_boundaries
from openpyxl.workbook import Workbook as WorkbookType
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill("solid", fgColor="1D4ED8")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def emit(payload: object, destination: str | None = None) -> None:
    text = json.dumps(payload, ensure_ascii=False, indent=2, default=str)
    if destination:
        path = Path(destination)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(f"{text}\n", encoding="utf-8")
    else:
        print(text)


def read_spec(path: str) -> dict[str, object]:
    value = json.loads(Path(path).read_text(encoding="utf-8"))
    if not isinstance(value, dict):
        raise TypeError("workbook specification must be a JSON object")
    return value


def style_header(worksheet: Worksheet, row_number: int = 1) -> None:
    for cell in worksheet[row_number]:
        if cell.value is None:
            continue
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def apply_number_formats(worksheet: Worksheet, formats: object) -> None:
    if formats is None:
        return
    if not isinstance(formats, dict):
        raise TypeError("numberFormats must be an object mapping ranges to formats")
    for cell_range, number_format in formats.items():
        for row in worksheet[str(cell_range)]:
            for cell in row:
                cell.number_format = str(number_format)


def apply_column_widths(worksheet: Worksheet, widths: object) -> None:
    if widths is None:
        return
    if not isinstance(widths, dict):
        raise TypeError("columnWidths must be an object")
    for column, width in widths.items():
        worksheet.column_dimensions[str(column)].width = float(width)


def reference_from_range(worksheet: Worksheet, expression: str) -> Reference:
    min_col, min_row, max_col, max_row = range_boundaries(expression)
    return Reference(
        worksheet,
        min_col=min_col,
        min_row=min_row,
        max_col=max_col,
        max_row=max_row,
    )


def add_charts(worksheet: Worksheet, charts: object) -> None:
    if charts is None:
        return
    if not isinstance(charts, list):
        raise TypeError("charts must be an array")
    chart_types = {"bar": BarChart, "line": LineChart, "pie": PieChart}
    for index, chart_spec in enumerate(charts, start=1):
        if not isinstance(chart_spec, dict):
            raise TypeError("every chart must be an object")
        chart_type = str(chart_spec.get("type", "bar")).lower()
        chart_class = chart_types.get(chart_type)
        if not chart_class:
            raise ValueError(f"unsupported chart type: {chart_type}")
        data_range = str(chart_spec.get("dataRange", ""))
        if not data_range:
            raise ValueError("chart dataRange is required")
        chart = chart_class()
        chart.title = str(chart_spec.get("title", ""))
        chart.style = int(chart_spec.get("style", 10))
        chart.add_data(
            reference_from_range(worksheet, data_range),
            titles_from_data=bool(chart_spec.get("titlesFromData", True)),
        )
        categories_range = str(chart_spec.get("categoriesRange", "")).strip()
        if categories_range:
            chart.set_categories(reference_from_range(worksheet, categories_range))
        chart.height = float(chart_spec.get("height", 7.5))
        chart.width = float(chart_spec.get("width", 12.0))
        worksheet.add_chart(
            chart, str(chart_spec.get("anchor", f"H{2 + (index - 1) * 16}"))
        )


def add_table(worksheet: Worksheet, value: object, index: int) -> None:
    if not value:
        return
    if worksheet.max_row < 2 or worksheet.max_column < 1:
        return
    if isinstance(value, dict):
        name = str(value.get("name", f"Table{index}"))
        reference = str(value.get("ref", worksheet.dimensions))
    else:
        name = f"Table{index}"
        reference = worksheet.dimensions
    table = Table(displayName=name, ref=reference)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def populate_sheet(worksheet: Worksheet, spec: dict[str, object], index: int) -> None:
    rows = spec.get("rows", [])
    if not isinstance(rows, list):
        raise TypeError("sheet rows must be an array")
    for row in rows:
        if not isinstance(row, list):
            raise TypeError("every worksheet row must be an array")
        worksheet.append(row)
    if bool(spec.get("styledHeader", True)) and rows:
        style_header(worksheet)
    freeze = str(spec.get("freezePanes", "")).strip()
    if freeze:
        worksheet.freeze_panes = freeze
    auto_filter = str(spec.get("autoFilter", "")).strip()
    if auto_filter:
        worksheet.auto_filter.ref = auto_filter
    for expression in spec.get("mergedCells", []):
        worksheet.merge_cells(str(expression))
    apply_column_widths(worksheet, spec.get("columnWidths"))
    apply_number_formats(worksheet, spec.get("numberFormats"))
    add_table(worksheet, spec.get("table"), index)
    add_charts(worksheet, spec.get("charts"))


def command_create(args: argparse.Namespace) -> None:
    spec = read_spec(args.spec)
    workbook = Workbook()
    workbook.properties.title = str(spec.get("title", ""))
    workbook.properties.creator = str(spec.get("author", "Wecode AI"))
    sheets = spec.get("sheets", [])
    if not isinstance(sheets, list) or not sheets:
        raise ValueError("sheets must be a non-empty array")
    for index, sheet_spec in enumerate(sheets, start=1):
        if not isinstance(sheet_spec, dict):
            raise TypeError("every sheet must be an object")
        title = str(sheet_spec.get("name", f"Sheet{index}"))
        worksheet = workbook.active if index == 1 else workbook.create_sheet()
        worksheet.title = title
        populate_sheet(worksheet, sheet_spec, index)
    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    emit(
        {
            "status": "created",
            "path": str(output.resolve()),
            "sheets": workbook.sheetnames,
        }
    )


def sheet_record(worksheet: Worksheet, max_cells: int) -> dict[str, object]:
    values: list[list[object]] = []
    consumed = 0
    truncated = False
    for row in worksheet.iter_rows():
        if consumed + len(row) > max_cells:
            truncated = True
            break
        values.append([cell.value for cell in row])
        consumed += len(row)
    formulas = [
        {"cell": cell.coordinate, "formula": cell.value}
        for row in worksheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]
    return {
        "name": worksheet.title,
        "dimensions": worksheet.dimensions,
        "maxRow": worksheet.max_row,
        "maxColumn": worksheet.max_column,
        "freezePanes": str(worksheet.freeze_panes or ""),
        "mergedCells": [str(value) for value in worksheet.merged_cells.ranges],
        "tables": sorted(worksheet.tables),
        "formulas": formulas,
        "values": values,
        "truncated": truncated,
    }


def command_inspect(args: argparse.Namespace) -> None:
    source = Path(args.input)
    workbook = load_workbook(source, data_only=False, read_only=False)
    emit(
        {
            "path": str(source.resolve()),
            "properties": {
                "title": workbook.properties.title or "",
                "creator": workbook.properties.creator or "",
            },
            "sheets": [
                sheet_record(sheet, args.max_cells) for sheet in workbook.worksheets
            ],
        },
        args.output,
    )


def command_set_cell(args: argparse.Namespace) -> None:
    workbook = load_workbook(args.input)
    if args.sheet not in workbook.sheetnames:
        raise ValueError(f"worksheet does not exist: {args.sheet}")
    if args.formula is not None:
        value: object = (
            args.formula if args.formula.startswith("=") else f"={args.formula}"
        )
    else:
        value = json.loads(args.value_json)
    workbook[args.sheet][args.cell] = value
    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    emit(
        {
            "status": "updated",
            "path": str(output.resolve()),
            "sheet": args.sheet,
            "cell": args.cell,
            "value": value,
        }
    )


def command_export_csv(args: argparse.Namespace) -> None:
    workbook = load_workbook(args.input, data_only=args.values_only, read_only=True)
    if args.sheet not in workbook.sheetnames:
        raise ValueError(f"worksheet does not exist: {args.sheet}")
    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.writer(stream)
        for row in workbook[args.sheet].iter_rows(values_only=True):
            writer.writerow(["" if value is None else value for value in row])
    emit({"status": "exported", "path": str(output.resolve()), "sheet": args.sheet})


def dependency_roots() -> list[Path]:
    configured = os.environ.get("CODEX_WORKSPACE_DEPENDENCIES", "").strip()
    if not configured:
        return []
    root = Path(configured).expanduser()
    return [root, root / "dependencies"]


def find_binary(name: str) -> str | None:
    resolved = shutil.which(name)
    if resolved:
        return resolved
    for root in dependency_roots():
        for candidate in (
            root / "bin" / "override" / name,
            root / "bin" / "fallback" / name,
            root / "native" / "poppler" / "bin" / name,
            root / "native" / "poppler" / "poppler" / "bin" / name,
        ):
            if candidate.is_file():
                return str(candidate)
    return None


def libreoffice_environment() -> dict[str, str]:
    environment = os.environ.copy()
    if environment.get("FONTCONFIG_FILE"):
        return environment
    for root in dependency_roots():
        configs = sorted(
            root.glob("native/libreoffice-headless/**/Resources/fontconfig/fonts.conf")
        )
        if configs:
            environment["FONTCONFIG_FILE"] = str(configs[0])
            environment["FONTCONFIG_PATH"] = str(configs[0].parent)
            break
    return environment


def command_render(args: argparse.Namespace) -> None:
    source = Path(args.input).resolve()
    output_dir = Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    office = find_binary("soffice") or find_binary("libreoffice")
    if not office:
        raise RuntimeError("LibreOffice is required to render XLSX files")
    subprocess.run(
        [
            office,
            "--headless",
            "--convert-to",
            "pdf",
            "--outdir",
            str(output_dir),
            str(source),
        ],
        check=True,
        capture_output=True,
        text=True,
        env=libreoffice_environment(),
    )
    pdf_path = output_dir / f"{source.stem}.pdf"
    if not pdf_path.is_file():
        raise RuntimeError("LibreOffice did not produce the expected PDF")
    images: list[str] = []
    if args.images:
        pdftoppm = find_binary("pdftoppm")
        if not pdftoppm:
            raise RuntimeError("pdftoppm is required when --images is requested")
        prefix = output_dir / source.stem
        subprocess.run(
            [pdftoppm, "-png", "-r", str(args.dpi), str(pdf_path), str(prefix)],
            check=True,
        )
        images = [str(path) for path in sorted(output_dir.glob(f"{source.stem}-*.png"))]
    emit({"pdf": str(pdf_path), "images": images})


def command_validate(args: argparse.Namespace) -> None:
    source = Path(args.input)
    issues: list[str] = []
    warnings: list[str] = []
    try:
        with zipfile.ZipFile(source) as archive:
            corrupt = archive.testzip()
            if corrupt:
                issues.append(f"corrupt ZIP entry: {corrupt}")
            names = set(archive.namelist())
            for required in ("[Content_Types].xml", "xl/workbook.xml"):
                if required not in names:
                    issues.append(f"missing XLSX part: {required}")
        workbook: WorkbookType = load_workbook(source, data_only=False)
        if not workbook.sheetnames:
            issues.append("workbook has no worksheets")
        for worksheet in workbook.worksheets:
            if worksheet.max_row > 1_048_576 or worksheet.max_column > 16_384:
                issues.append(f"{worksheet.title}: used range exceeds XLSX limits")
            formula_count = sum(
                1
                for row in worksheet.iter_rows()
                for cell in row
                if isinstance(cell.value, str) and cell.value.startswith("=")
            )
            if formula_count:
                warnings.append(
                    f"{worksheet.title}: {formula_count} formulas require recalculation by a spreadsheet app"
                )
    except Exception as error:  # noqa: BLE001 - validation reports format failures
        issues.append(str(error))
    emit(
        {
            "valid": not issues,
            "issues": issues,
            "warnings": warnings,
            "path": str(source.resolve()),
        }
    )
    if issues:
        raise SystemExit(1)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    subparsers = parser.add_subparsers(dest="command", required=True)
    create = subparsers.add_parser("create")
    create.add_argument("--spec", required=True)
    create.add_argument("--output", required=True)
    create.set_defaults(handler=command_create)
    inspect = subparsers.add_parser("inspect")
    inspect.add_argument("--input", required=True)
    inspect.add_argument("--output")
    inspect.add_argument("--max-cells", type=int, default=1000)
    inspect.set_defaults(handler=command_inspect)
    set_cell = subparsers.add_parser("set-cell")
    set_cell.add_argument("--input", required=True)
    set_cell.add_argument("--sheet", required=True)
    set_cell.add_argument("--cell", required=True)
    value_group = set_cell.add_mutually_exclusive_group(required=True)
    value_group.add_argument("--value-json")
    value_group.add_argument("--formula")
    set_cell.add_argument("--output", required=True)
    set_cell.set_defaults(handler=command_set_cell)
    export_csv = subparsers.add_parser("export-csv")
    export_csv.add_argument("--input", required=True)
    export_csv.add_argument("--sheet", required=True)
    export_csv.add_argument("--output", required=True)
    export_csv.add_argument("--values-only", action="store_true")
    export_csv.set_defaults(handler=command_export_csv)
    render = subparsers.add_parser("render")
    render.add_argument("--input", required=True)
    render.add_argument("--output-dir", required=True)
    render.add_argument("--images", action="store_true")
    render.add_argument("--dpi", type=int, default=144)
    render.set_defaults(handler=command_render)
    validate = subparsers.add_parser("validate")
    validate.add_argument("--input", required=True)
    validate.set_defaults(handler=command_validate)
    return parser


def main() -> None:
    args = build_parser().parse_args()
    try:
        args.handler(args)
    except (OSError, ValueError, RuntimeError, subprocess.CalledProcessError) as error:
        print(f"spreadsheets: {error}", file=sys.stderr)
        raise SystemExit(2) from error


if __name__ == "__main__":
    main()
